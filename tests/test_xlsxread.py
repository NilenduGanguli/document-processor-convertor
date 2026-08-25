"""Tests for :mod:`dpc.xlsxread`.

Every workbook is built in-test with openpyxl — one exception: the last test runs an actual
corpus file end-to-end through ``read_xlsx`` + ``to_pmd`` (skipped where the corpus is not
present).

On formulas: openpyxl-authored workbooks are never calculated, so they carry **no cached
formula results**. ``read_only=True, data_only=True`` therefore reads their formula cells
as ``None`` -> ``""`` — asserted below as documented behaviour, not worked around. A file
last saved by a real spreadsheet application would carry the cache and yield the value.
"""
from __future__ import annotations

import datetime
import io
import pathlib

import openpyxl
import pytest

from dpc import xlsxread
from dpc.emitter import to_pmd
from dpc.models import Zone
from dpc.xlsxread import MAX_COLS, MAX_ROWS, read_xlsx

CORPUS = pathlib.Path(
    "~/document-classification-extraction/corpus/ca/ca_isc_register.xlsx"
).expanduser()


def workbook_bytes(workbook: openpyxl.Workbook) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def basic_workbook() -> bytes:
    """Two sheets: a small register and a one-cell notes sheet."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Register"
    sheet.append(["Name", "Qty", "When"])
    sheet.append(["widget", 3, datetime.datetime(2024, 1, 5)])  # noqa: DTZ001 — Excel dates are naive
    notes = workbook.create_sheet("Notes")
    notes["A1"] = "remember"
    return workbook_bytes(workbook)


# ---------------------------------------------------------------------------
# Structure: pages, sheet-name headings, tables
# ---------------------------------------------------------------------------
def test_one_page_per_sheet_with_no_fake_sizes() -> None:
    view = read_xlsx(basic_workbook())
    assert [p.page for p in view.pages] == [1, 2]
    assert all(p.width == 0.0 and p.height == 0.0 for p in view.pages)


def test_sheet_name_is_a_heading_block_with_role_sheet_and_no_bbox() -> None:
    view = read_xlsx(basic_workbook())
    assert [b.text for b in view.blocks] == ["Register", "Notes"]
    assert all(b.zone is Zone.heading for b in view.blocks)
    assert all(b.role == "sheet" for b in view.blocks)
    assert all(b.bbox is None for b in view.blocks)  # no geometry -> no invented anchor
    assert [b.page for b in view.blocks] == [1, 2]


def test_one_table_per_sheet_on_its_page_and_no_table_zone_blocks() -> None:
    view = read_xlsx(basic_workbook())
    assert [t.page for t in view.tables] == [1, 2]
    assert all(t.bbox is None for t in view.tables)
    # Table text must not also exist as blocks — the only blocks are the sheet names.
    assert all(b.role == "sheet" for b in view.blocks)


def test_first_row_is_header_by_convention() -> None:
    view = read_xlsx(basic_workbook())
    table = view.tables[0]
    assert all(c.is_header for c in table.cells if c.row == 0)
    assert not any(c.is_header for c in table.cells if c.row > 0)


def test_empty_sheet_gets_page_and_heading_but_no_table() -> None:
    workbook = openpyxl.Workbook()
    workbook.active.title = "Blank"
    view = read_xlsx(workbook_bytes(workbook))
    assert [p.page for p in view.pages] == [1]
    assert [b.text for b in view.blocks] == ["Blank"]
    assert view.tables == []


def test_not_a_workbook_raises_valueerror_without_content() -> None:
    with pytest.raises(ValueError, match="cannot open XLSX"):
        read_xlsx(b"definitely not a zip")


# ---------------------------------------------------------------------------
# Cell values
# ---------------------------------------------------------------------------
def test_values_render_without_noise() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["text", 1.0, 2.5, 7, None])
    sheet.append([datetime.datetime(2024, 1, 5), datetime.datetime(2024, 1, 5, 6, 0),  # noqa: DTZ001 — Excel dates are naive
                  True, False, ""])
    view = read_xlsx(workbook_bytes(workbook))
    grid = view.tables[0].grid()
    # Column 5 (None / "") is a fully-empty trailing column, so the trim removes it.
    assert grid[0] == ["text", "1", "2.5", "7"]
    assert grid[1] == ["2024-01-05", "2024-01-05T06:00:00", "TRUE", "FALSE"]


def test_uncached_formula_reads_as_blank() -> None:
    """data_only returns the *cached* result; an openpyxl-authored file has none -> ""."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = 2
    sheet["A2"] = "=SUM(A1,A1)"
    sheet["B2"] = "anchor-col"  # keeps the formula cell inside the trimmed grid
    view = read_xlsx(workbook_bytes(workbook))
    assert view.tables[0].grid()[1] == ["", "anchor-col"]


# ---------------------------------------------------------------------------
# Merged cells
# ---------------------------------------------------------------------------
def test_merge_lands_top_left_with_spans_and_no_covered_cells() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Name", "Qty"])
    sheet.append(["widget", 3])
    sheet.merge_cells("A3:B4")
    sheet["A3"] = "merged"
    view = read_xlsx(workbook_bytes(workbook))
    table = view.tables[0]
    # The merge reaches row 4: a value-bearing merge keeps the rows/cols it spans over.
    assert (table.row_count, table.col_count) == (4, 2)
    merged = {(c.row, c.col): c for c in table.cells}
    assert merged[(2, 0)].text == "merged"
    assert (merged[(2, 0)].row_span, merged[(2, 0)].col_span) == (2, 2)
    assert (2, 1) not in merged and (3, 0) not in merged and (3, 1) not in merged
    assert len(table.cells) == 5  # 2x2 grid above + one merged cell


# ---------------------------------------------------------------------------
# Trimming and caps
# ---------------------------------------------------------------------------
def test_trailing_empty_rows_and_columns_are_trimmed_interior_kept() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "a"
    sheet["A3"] = "b"          # interior empty row 2 must survive
    sheet["E7"] = None         # grows the dimension with nothing in it
    view = read_xlsx(workbook_bytes(workbook))
    table = view.tables[0]
    assert (table.row_count, table.col_count) == (3, 1)
    assert table.grid() == [["a"], [""], ["b"]]
    assert "truncated" not in view.raw  # trimming is not truncation


def test_row_cap_bites_and_is_recorded() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "head"
    sheet.cell(row=MAX_ROWS + 2, column=1, value="beyond the cap")
    view = read_xlsx(workbook_bytes(workbook))
    assert view.raw["truncated"] is True
    assert view.tables[0].row_count == 1  # kept rows below the cap were empty -> trimmed


def test_col_cap_bites_and_is_recorded() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "a"
    sheet.cell(row=1, column=MAX_COLS + 6, value="beyond the cap")
    view = read_xlsx(workbook_bytes(workbook))
    assert view.raw["truncated"] is True
    assert view.tables[0].col_count == 1


def test_dimension_only_padding_beyond_caps_is_not_truncation() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "a"
    sheet.cell(row=MAX_ROWS + 50, column=1, value=None)  # dimension, no content
    view = read_xlsx(workbook_bytes(workbook))
    assert "truncated" not in view.raw


# ---------------------------------------------------------------------------
# End-to-end: an actual corpus workbook through read_xlsx + to_pmd
# ---------------------------------------------------------------------------
@pytest.mark.skipif(not CORPUS.exists(), reason="corpus workbook not present")
def test_corpus_register_end_to_end() -> None:
    view = read_xlsx(CORPUS.read_bytes())
    assert view.pages and view.blocks and view.tables
    assert all(b.zone is not Zone.table for b in view.blocks)

    pmd = to_pmd(
        view,
        source="document",
        provider=xlsxread.PROVIDER_OPENPYXL,
        generated="2026-01-01T00:00:00+00:00",
    )
    lines = pmd.splitlines()
    # A GFM table: a pipe row plus its delimiter row.
    assert any(line.startswith("| ") and line.endswith(" |") for line in lines)
    assert any(line.startswith("|") and set(line) <= {"|", "-", " "} and "---" in line
               for line in lines)
    # The sheet name appears as a markdown heading.
    assert f"## {view.blocks[0].text}" in lines
    # Nothing in the view carried geometry, so nothing may claim any: no anchors at all.
    assert not any(line.startswith("<!-- @") for line in lines)
